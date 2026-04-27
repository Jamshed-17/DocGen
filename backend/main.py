# document-generator/backend/main.py
import io
import os
import zipfile
import csv
import re
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from PIL import Image, ImageDraw, ImageFont
import img2pdf
from openpyxl import load_workbook
import textwrap 
from fastapi.staticfiles import StaticFiles

# --- Константы и Глобальное Состояние ---
UPLOAD_DIR = "uploads"
CURRENT_TEMPLATE_PATH = os.path.join(UPLOAD_DIR, "current_template.png")
CURRENT_DATA_PATH = os.path.join(UPLOAD_DIR, "data.uploaded")
GENERATION_DATA = [] # Монтируем папку шрифтов, чтобы они были доступны по URL /fonts/имя_файла

# Настройка шрифтов
FONTS_DIR = "data/fonts"
DEFAULT_FONT_NAME = "Arial"
AVAILABLE_FONTS = {}
# -------------------------------------------

app = FastAPI()
# Монтируем папку шрифтов, чтобы они были доступны по URL /fonts/имя_файла
app.mount("/fonts", StaticFiles(directory=FONTS_DIR), name="fonts")


# --- Настройка CORS и Инициализация ---
origins = ["http://localhost", "http://localhost:3000"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """Сканирование папки шрифтов при запуске."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(FONTS_DIR, exist_ok=True)
    
    # Заполнение доступных шрифтов
    for filename in os.listdir(FONTS_DIR):
        if filename.lower().endswith(('.ttf', '.otf')):
            name = os.path.splitext(filename)[0]
            AVAILABLE_FONTS[name] = os.path.join(FONTS_DIR, filename)

# ---------------------------------------------------------
# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
# ---------------------------------------------------------

def apply_text_case(text: str, case_type: str) -> str:
    """Применяет преобразование регистра к тексту."""
    if not text:
        return ""

    if case_type == 'upper':
        return text.upper()
    elif case_type == 'lower':
        return text.lower()
    elif case_type == 'title':
        # Каждое Слово С Большой
        return text.title()
    elif case_type == 'sentence':
        # Большая буква с начала предложения
        # Разделение по знакам препинания с сохранением разделителя
        sentences = re.split(r'([.!?])\s*', text.strip())
        result = []
        is_new_sentence = True
        for part in sentences:
            if not part:
                continue
            
            # Убираем пробелы, делаем первую букву заглавной
            if is_new_sentence and part.strip():
                 result.append(part.strip().capitalize())
                 is_new_sentence = False
            elif part in ['.', '!', '?']:
                 # Это разделитель, следующий элемент - новое предложение
                 result.append(part)
                 is_new_sentence = True
            elif part.strip():
                 # Не-разделитель внутри предложения (слово)
                 if result and result[-1] in ['.', '!', '?']:
                     # Если предыдущий элемент был знаком препинания, начинаем с заглавной
                      result.append(part.strip().capitalize())
                      is_new_sentence = False
                 else:
                      result.append(part.strip().lower()) # Внутри предложения все строчные
                      is_new_sentence = False
        
        # Собираем все части, вставляя пробелы там, где они были бы естественны
        final_text = ""
        for part in result:
             if part in ['.', '!', '?']:
                  final_text += part
             elif final_text and final_text[-1] not in ['.', '!', '?'] and final_text[-1] != ' ':
                  final_text += ' ' + part
             else:
                  final_text += part

        return final_text.strip()
    else:
        return text

def wrap_text_to_width(text: str, font: ImageFont.ImageFont, block_width: int) -> str:
    """Реализует перенос текста по словам на заданную ширину."""
    if block_width <= 0:
        return text
    
    try:
        current_line = []
        wrapped_lines = []
        current_width = 0
        
        words = text.split(' ')
        
        for word in words:
            word_width = font.getlength(word)
            
            # Проверяем, поместится ли слово в текущую строку
            # Учитываем пробел только если строка не пуста
            space_width = font.getlength(' ') if current_line else 0
            
            if current_width + word_width + space_width <= block_width:
                current_line.append(word)
                current_width += word_width + space_width
            else:
                if current_line:
                    wrapped_lines.append(' '.join(current_line))
                
                # Если слово само по себе помещается, оно начинает новую строку
                if word_width <= block_width:
                    current_line = [word]
                    current_width = word_width
                else:
                    # Слово длиннее ширины блока (перенос по символам не поддерживается PIL, оставляем как есть)
                    wrapped_lines.append(word)
                    current_line = []
                    current_width = 0
                
        if current_line:
            wrapped_lines.append(' '.join(current_line))
            
        return "\n".join(wrapped_lines)
        
    except Exception:
        return text

def parse_uploaded_data(file_bytes: bytes, filename: str):
    """Определяет формат файла, парсит его и реализует логику пропуска/остановки."""
    
    filename_lower = filename.lower()
    data = []
    headers = []

    if filename_lower.endswith(('.xlsx', '.xls')):
        # --- ЛОГИКА ДЛЯ EXCEL (XLSX) ---
        try:
            workbook = load_workbook(io.BytesIO(file_bytes))
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1] if cell.value is not None]
            
            empty_row_count = 0
            for row_index in range(2, sheet.max_row + 1):
                record = {}
                row_is_empty = True
                
                for col_index, header in enumerate(headers):
                    value = sheet.cell(row=row_index, column=col_index + 1).value
                    record[header] = value
                    
                    if value is not None and str(value).strip() != "":
                        row_is_empty = False
                
                # ЛОГИКА ПРОПУСКА И ОСТАНОВКИ (7 пустых строк)
                if row_is_empty:
                    empty_row_count += 1
                    if empty_row_count >= 7: 
                        break
                    continue
                else:
                    empty_row_count = 0 
                    data.append(record)
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Ошибка при парсинге Excel: {e}. Проверьте формат файла.")
        
    elif filename_lower.endswith('.csv'):
        # --- ЛОГИКА ДЛЯ CSV ---
        try:
            csv_data = file_bytes.decode('utf-8-sig') 
        except:
            csv_data = file_bytes.decode('cp1251', errors='ignore')
            
        reader = csv.DictReader(io.StringIO(csv_data))
        headers = [h.strip() for h in reader.fieldnames if h is not None and h.strip()]
        
        empty_row_count = 0
        for row in reader:
            record = {}
            row_is_empty = True

            for key, value in row.items():
                if key is not None:
                     record[key.strip()] = value
                if value is not None and str(value).strip() != "":
                     row_is_empty = False
            
            # ЛОГИКА ПРОПУСКА И ОСТАНОВКИ (7 пустых строк)
            if row_is_empty:
                empty_row_count += 1
                if empty_row_count >= 7:
                    break
                continue
            else:
                empty_row_count = 0
                data.append(record)
    
    else:
        raise HTTPException(status_code=400, detail="Поддерживается только Excel (xlsx/xls) и CSV.")

    return data, headers
# ---------------------------------------------------------


@app.get("/")
def read_root():
    return {"message": "Backend is running and ready for generation!"}

@app.get("/available-fonts/")
def get_available_fonts():
    fonts_info = []
    if os.path.exists(FONTS_DIR):
        for filename in os.listdir(FONTS_DIR):
            if filename.lower().endswith(('.ttf', '.otf')):
                # Возвращаем объект: имя для CSS и имя файла для URL
                fonts_info.append({
                    "name": os.path.splitext(filename)[0],
                    "file": filename
                })
    return fonts_info

# --- МАРШРУТЫ ЗАГРУЗКИ (без изменений) ---

@app.post("/upload-template/")
async def upload_template(file: UploadFile = File(...)):
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Файл должен быть изображением.")
    
    try:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        img = Image.open(io.BytesIO(await file.read()))
        img.save(CURRENT_TEMPLATE_PATH)
        width, height = img.size
        return {"status": "success", "filename": file.filename, "width": width, "height": height}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка сервера при обработке изображения: {e}")

@app.post("/upload-data/")
async def upload_data(file: UploadFile = File(...)):
    global GENERATION_DATA
    
    file_bytes = await file.read()
    
    try:
        GENERATION_DATA, headers = parse_uploaded_data(file_bytes, file.filename)
    except HTTPException as e:
         raise e
    except Exception as e:
         raise HTTPException(status_code=500, detail=f"Критическая ошибка при обработке данных: {e}")
    
    with open(CURRENT_DATA_PATH, "wb") as f:
        f.write(file_bytes)
    
    if not GENERATION_DATA:
        raise HTTPException(status_code=400, detail="Файл данных пуст или содержит только заголовки.")
    
    return {"status": "success", "keys": headers, "count": len(GENERATION_DATA)}

@app.post("/upload-font/")
async def upload_font(file: UploadFile = File(...)):
    """Загрузка нового шрифта .ttf/.otf."""
    if not file.filename.lower().endswith(('.ttf', '.otf')):
        raise HTTPException(status_code=400, detail="Поддерживаются только шрифты TTF и OTF.")
    
    font_name = os.path.splitext(file.filename)[0]
    font_path = os.path.join(FONTS_DIR, file.filename)
    
    try:
        # Проверка, что файл еще не существует
        if os.path.exists(font_path):
             raise HTTPException(status_code=409, detail=f"Шрифт с именем '{font_name}' уже существует.")

        with open(font_path, "wb") as f:
            f.write(await file.read())
        
        # Обновляем глобальное состояние
        AVAILABLE_FONTS[font_name] = font_path
        
        return {"status": "success", "font_name": font_name, "message": "Шрифт успешно загружен."}
    except HTTPException as e:
         raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при сохранении шрифта: {e}")

# --- МАРШРУТ ГЕНЕРАЦИИ ---

@app.post("/generate/")
async def generate_documents(
    config_data: dict
):
    global GENERATION_DATA
    try:
        template_config = config_data.get("template_config", [])
        
        options = config_data.get("generation_options", {})
        file_name_template = options.get("fileNameTemplate", "document_{INDEX}")
        output_mode = options.get("outputMode", "zip") 

        # 1. Проверка
        if not os.path.exists(CURRENT_TEMPLATE_PATH):
            raise FileNotFoundError("Шаблон не загружен. Загрузите изображение.")
        if not GENERATION_DATA:
            raise HTTPException(status_code=400, detail="Данные для генерации не загружены.")

        img_template = Image.open(CURRENT_TEMPLATE_PATH)
        img_template_rgba = img_template.convert("RGBA") 
        
        zip_buffer = io.BytesIO()
        image_buffers_for_pdf = [] 
        pdf_buffers = [] 
        
        # 2. Цикл по каждой строке данных
        for i, record in enumerate(GENERATION_DATA):
            
            img = img_template_rgba.copy() 
            draw = ImageDraw.Draw(img)
            template_width, template_height = img.size
            
            # 3. Обрабатываем каждый блок разметки
            for block in template_config:
                
                # 3.1. Условный рендеринг и замена ключей
                condition_key = block.get("conditionKey")
                if_empty_text = block.get("ifEmptyText", "")
                raw_template = block.get("template", "")
                text_to_write = ""

                is_condition_met = True
                
                if condition_key:
                    condition_value = record.get(condition_key)
                    if condition_value is None or str(condition_value).strip() == "":
                         is_condition_met = False
                         
                
                if not is_condition_met and condition_key:
                     text_to_write = if_empty_text
                
                else:
                    keys_found = re.findall(r'\{([^{}]+)\}', raw_template)
                    text_to_write = raw_template
                    for key in keys_found:
                        value = record.get(key.strip())
                        text_to_write = text_to_write.replace(f"{{{key}}}", str(value) if value is not None else "")
                        
                if not text_to_write.strip() and not if_empty_text:
                    continue

                # NEW: Применяем преобразование регистра
                text_case = block.get("textCase", "none")
                text_to_write = apply_text_case(text_to_write, text_case)
                
                
                # 3.2. Настройка шрифта и координат
                font_name = block.get("fontName", DEFAULT_FONT_NAME)
                font_path = AVAILABLE_FONTS.get(font_name)
                font_size = block.get("fontSize", 24)
                
                if not isinstance(font_size, int) or font_size <= 0:
                    font_size = 24
                
                # Загрузка начального шрифта
                try:
                    if font_path:
                        font = ImageFont.truetype(font_path, size=font_size)
                    else:
                        font = ImageFont.load_default(size=font_size)
                except Exception:
                    font = ImageFont.load_default(size=font_size)

                # Координаты левого верхнего угла, ширина и высота блока в пикселях
                x_pixel = int(block.get("x_percent", 0) * template_width / 100)
                y_pixel = int(block.get("y_percent", 0) * template_height / 100)
                block_width = int(block.get("width_percent", 0) * template_width / 100) 
                height_pixel = int(block.get("height_percent", 0) * template_height / 100)
                
                alignment = block.get("alignment", "left")
                font_color = block.get("color", "#000000")
                
                # NEW: Параметры Shrink to Fit
                shrink_to_fit = block.get("shrinkToFit", False)
                current_font_size = font_size
                min_font_size = 8 # Минимальный размер шрифта

                # 3.3. **Цикл Shrink to Fit (Если включен)**
                wrapped_text = ""
                
                if shrink_to_fit and block_width > 0 and height_pixel > 0:
                    
                    while current_font_size >= min_font_size:
                        try:
                            # Пересоздаем шрифт
                            font = ImageFont.truetype(font_path, size=current_font_size)
                        except Exception:
                            font = ImageFont.load_default(size=current_font_size)
                        
                        # Перенос текста с текущим размером
                        wrapped_text = wrap_text_to_width(text_to_write, font, block_width)
                        
                        # Проверка высоты
                        text_height = 0
                        if wrapped_text.strip():
                            try:
                                # PIL getsize is deprecated, use getlength/textbbox
                                bbox = draw.textbbox((0, 0), wrapped_text, font=font, anchor="lt", align=alignment)
                                text_height = bbox[3] - bbox[1] 
                            except Exception:
                                line_count = wrapped_text.count('\n') + 1
                                text_height = current_font_size * line_count * 1.2
                        
                        # Проверка на вмещение
                        if text_height <= height_pixel:
                            break # Поместилось! Выходим из цикла.

                        current_font_size -= 1 # Уменьшаем размер и пробуем снова

                    if current_font_size < min_font_size:
                        # Если не влез даже минимальный, берем минимальный размер и текст
                        try:
                            font = ImageFont.truetype(font_path, size=min_font_size)
                        except Exception:
                            font = ImageFont.load_default(size=min_font_size)
                        wrapped_text = wrap_text_to_width(text_to_write, font, block_width)
                        
                else:
                    # Оригинальная логика (без уменьшения)
                    wrapped_text = wrap_text_to_width(text_to_write, font, block_width)

                
                # 3.4. **Вертикальное центрирование**
                text_height = 0
                y_final = y_pixel
                
                if wrapped_text.strip():
                    try:
                        # Получаем ограничивающую рамку текста (используем финальный font)
                        bbox = draw.textbbox((0, 0), wrapped_text, font=font, anchor="lt", align=alignment)
                        text_height = bbox[3] - bbox[1] 
                    except Exception:
                        line_count = wrapped_text.count('\n') + 1
                        # Используем финальный font size
                        final_font_size = current_font_size if shrink_to_fit else font_size 
                        text_height = final_font_size * line_count * 1.2
                        
                    # Вычисляем стартовую Y-координату для центрирования
                    if height_pixel > 0 and text_height > 0:
                        y_start_offset = (height_pixel - text_height) // 2
                        y_final = y_pixel + max(0, y_start_offset) 
                    else:
                        y_final = y_pixel
                
                # 3.5. Рисование фона
                background_color = block.get("backgroundColor")
                
                if background_color and block_width > 0 and height_pixel > 0:
                    x1 = x_pixel
                    y1 = y_pixel
                    x2 = x1 + block_width
                    y2 = y1 + height_pixel

                    draw.rectangle([x1, y1, x2, y2], fill=background_color)
                
                
                # 3.6. Вставка текста
                lines = wrapped_text.split('\n')

                # Рассчитываем вертикальный межстрочный интервал (примерно 1.2 от размера шрифта)
                # Или используем textbbox для точного замера высоты одной строки
                line_spacing = font.size * 1.2 

                for j, line in enumerate(lines):
                    # Вычисляем Y для текущей строки
                    current_y = y_final + (j * line_spacing)
                    
                    # Вычисляем X в зависимости от выравнивания
                    line_bbox = draw.textbbox((0, 0), line, font=font)
                    line_width = line_bbox[2] - line_bbox[0]
                    
                    if alignment == "center":
                        x_start = x_pixel + (block_width - line_width) // 2
                    elif alignment == "right":
                        x_start = x_pixel + block_width - line_width
                    else:
                        x_start = x_pixel

                    # Рисуем каждую строку отдельно
                    draw.text(
                        (x_start, current_y), 
                        line,
                        fill=font_color,
                        font=font
                    )
            
            # 4. Сохранение изображения в буфер (имя файла, режим вывода)
            # ... (логика формирования имени файла - без изменений)
            temp_name = file_name_template
            keys_found = re.findall(r'\{([^{}]+)\}', temp_name)
            for key in keys_found:
                key_name = key.strip()
                if key_name.upper() == "INDEX":
                    value = str(i + 1).zfill(3) 
                else:
                    value = record.get(key_name)
                temp_name = temp_name.replace(f"{{{key_name}}}", str(value) if value is not None else "")
                
            base_name = re.sub(r'[\\/:*?"<>|]+', '_', temp_name).strip()
            if not base_name:
                base_name = str(i + 1).zfill(3)
            
            filename = f"{base_name}.pdf"


            # 5. Сбор буферов
            img_buffer = io.BytesIO()
            img.save(img_buffer, format="PNG")
            img_buffer.seek(0)
            
            if output_mode == 'single_pdf':
                 image_buffers_for_pdf.append(img_buffer.getvalue()) 
            else:
                pdf_bytes = img2pdf.convert(img_buffer.getvalue()) 
                pdf_buffers.append((filename, pdf_bytes))
                

        # --- После цикла: Выдача результата ---
        
        if output_mode == 'zip':
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for filename, pdf_bytes in pdf_buffers:
                    zip_file.writestr(filename, pdf_bytes)
                    
            zip_buffer.seek(0)
            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": "attachment; filename=documents.zip"}
            )
            
        elif output_mode == 'single_pdf':
            if not image_buffers_for_pdf:
                raise HTTPException(status_code=500, detail="Не удалось создать изображения для объединения.")
                
            merged_pdf_bytes = img2pdf.convert(image_buffers_for_pdf)
            
            pdf_buffer = io.BytesIO(merged_pdf_bytes)
            pdf_buffer.seek(0)
            
            return StreamingResponse(
                pdf_buffer,
                media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=merged_documents.pdf"}
            )
            
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        print(f"Критическая ошибка при генерации: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка сервера: {e}")